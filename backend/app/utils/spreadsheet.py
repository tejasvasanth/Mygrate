"""Parse uploaded spreadsheets (xlsx / csv) into rows of dicts, in memory."""
from __future__ import annotations

import csv
import io
from typing import Any

MAX_ROWS = 50_000


def parse_spreadsheet(filename: str, content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    """Returns (headers, rows). First row is treated as the header row."""
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return _parse_xlsx(content)
    if name.endswith((".csv", ".txt", ".tsv")):
        return _parse_csv(content, delimiter="\t" if name.endswith(".tsv") else ",")
    raise ValueError("Unsupported file type — upload .xlsx, .csv or .tsv")


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError("Workbook has no active sheet")
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        raise ValueError("Spreadsheet is empty") from None
    headers = [str(h).strip() if h is not None else f"column_{i+1}"
               for i, h in enumerate(header_row)]
    rows: list[dict[str, Any]] = []
    for raw in it:
        if all(v is None for v in raw):
            continue
        rows.append({headers[i]: raw[i] if i < len(raw) else None
                     for i in range(len(headers))})
        if len(rows) >= MAX_ROWS:
            break
    wb.close()
    return headers, rows


def _parse_csv(content: bytes, delimiter: str = ",") -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [h.strip() for h in (reader.fieldnames or [])]
    if not headers:
        raise ValueError("CSV has no header row")
    rows: list[dict[str, Any]] = []
    for r in reader:
        if all((v is None or str(v).strip() == "") for v in r.values()):
            continue
        rows.append({(k or "").strip(): _coerce(v) for k, v in r.items()})
        if len(rows) >= MAX_ROWS:
            break
    return headers, rows


def _coerce(v: Any) -> Any:
    """CSV gives strings only — coerce obvious numbers/booleans/nulls."""
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    low = s.lower()
    if low in ("true", "false"):
        return low == "true"
    if low in ("null", "none", "n/a"):
        return None
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s
